#!/usr/bin/env python3
"""
Convert ResAnalytics "Unit Availability" .xlsx files to CSV.

Keeps:
- the two-row table header, combined into one CSV row
- property data rows

Removes:
- report title/property/as-of lines above the header
- trailing Total row(s)

Reads all matching workbooks from data/raw/Unit_Availability and writes
data/csv/unit_availability/<identifier>.csv.
"""

from __future__ import annotations

import csv
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


DATA_DIR = Path(__file__).resolve().parent
INPUT_DIR = DATA_DIR / "raw" / "Unit_Availability"
OUTPUT_DIR = DATA_DIR / "csv" / "unit_availability"
FILENAME_PREFIX = "ResAnalytics_Unit_Availability_"


def csv_value(value):
    """Serialize Excel values without otherwise changing the data."""
    if isinstance(value, datetime):
        return value.strftime("%m/%d/%Y")
    if isinstance(value, date):
        return value.strftime("%m/%d/%Y")
    return value


def combined_header(ws, header_row: int):
    """Combine the two Excel header rows column by column."""
    rows = ws.iter_rows(
        min_row=header_row,
        max_row=header_row + 1,
        values_only=True,
    )
    first, second = rows
    return [
        " ".join(str(value) for value in values if value is not None)
        for values in zip(first, second)
    ]


def find_header_row(ws) -> int:
    """Find the first header row: Property, Name, Avg., ..."""
    for row in ws.iter_rows():
        if row[0].value == "Property" and row[1].value == "Name":
            return row[0].row
    raise ValueError("Could not find Unit Availability header row")


def convert_file(input_path: Path, output_path: Path) -> None:
    wb = load_workbook(input_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    header_row = find_header_row(ws)
    data_start_row = header_row + 2

    rows_to_write = [combined_header(ws, header_row)]

    # Keep property rows; omit blank rows and trailing Total row(s).
    for row in ws.iter_rows(min_row=data_start_row, values_only=True):
        if all(value is None for value in row):
            continue

        # In the supplied reports, the summary row is identified by Name=Total.
        if len(row) > 1 and row[1] == "Total":
            continue

        rows_to_write.append(row)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        for row in rows_to_write:
            writer.writerow(csv_value(value) for value in row)

    wb.close()
    print(f"Converted: {input_path} -> {output_path}")


def main() -> None:
    paths = sorted(INPUT_DIR.glob(f"{FILENAME_PREFIX}*.xlsx"))
    if not paths:
        raise SystemExit(f"No matching .xlsx files found in: {INPUT_DIR}")

    for path in paths:
        identifier = path.stem[len(FILENAME_PREFIX) :]
        convert_file(path, OUTPUT_DIR / f"{identifier}.csv")

    print(f"Done. Converted {len(paths)} Unit Availability file(s) to {OUTPUT_DIR}.")


if __name__ == "__main__":
    main()
