#!/usr/bin/env python3
"""
Convert ResAnalytics "Rent Roll with Lease Charges" .xlsx files to CSV.

Keeps:
- the two-row table header, combined into one CSV row
- Current/Notice/Vacant Residents section
- Future Residents/Applicants section
- all unit/resident/charge rows and blank separator rows inside the body

Removes:
- report title/property/as-of/month lines above the header
- Summary Groups and everything after it

Reads all matching workbooks from data/raw/Rent_Roll_with_Lease_Charges and
writes data/csv/rent_roll/<identifier>.csv.
"""

from __future__ import annotations

import csv
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


DATA_DIR = Path(__file__).resolve().parent
INPUT_DIR = DATA_DIR / "raw" / "Rent_Roll_with_Lease_Charges"
OUTPUT_DIR = DATA_DIR / "csv" / "rent_roll"
FILENAME_PREFIX = "ResAnalytics_Rent_Roll_with_Lease_Charges_"


def csv_value(value):
    """Serialize Excel values without otherwise changing the data."""
    if isinstance(value, datetime):
        return value.strftime("%m/%d/%Y")
    if isinstance(value, date):
        return value.strftime("%m/%d/%Y")
    return value


def combined_header(ws, header_row: int):
    """Combine the two Excel header rows column by column."""
    rows = ws.iter_rows(
        min_row=header_row,
        max_row=header_row + 1,
        values_only=True,
    )
    first, second = rows
    return [
        " ".join(str(value) for value in values if value is not None)
        for values in zip(first, second)
    ]


def find_header_row(ws) -> int:
    """Find the first header row: Unit, Unit Type, Unit, Resident, ..."""
    for row in ws.iter_rows():
        values = [cell.value for cell in row]
        if (
            len(values) >= 4
            and values[0] == "Unit"
            and values[1] == "Unit Type"
            and values[3] == "Resident"
        ):
            return row[0].row
    raise ValueError("Could not find Rent Roll header row")


def find_body_end_row(ws, start_row: int) -> int:
    """
    The main body ends immediately before the 'Summary Groups' section.
    """
    for row in ws.iter_rows(min_row=start_row):
        if row[0].value == "Summary Groups":
            return row[0].row - 1

    # Fallback: if no summary marker exists, keep through the final used row.
    return ws.max_row


def convert_file(input_path: Path, output_path: Path) -> None:
    wb = load_workbook(input_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    header_row = find_header_row(ws)
    body_end_row = find_body_end_row(ws, header_row + 2)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(combined_header(ws, header_row))

        # The two header rows were combined above; write only the body here.
        for row in ws.iter_rows(
            min_row=header_row + 3,
            max_row=body_end_row,
            values_only=True,
        ):
            writer.writerow(csv_value(value) for value in row)

    wb.close()
    print(f"Converted: {input_path} -> {output_path}")


def main() -> None:
    paths = sorted(INPUT_DIR.glob(f"{FILENAME_PREFIX}*.xlsx"))
    if not paths:
        raise SystemExit(f"No matching .xlsx files found in: {INPUT_DIR}")

    for path in paths:
        identifier = path.stem[len(FILENAME_PREFIX) :]
        convert_file(path, OUTPUT_DIR / f"{identifier}.csv")

    print(f"Done. Converted {len(paths)} Rent Roll file(s) to {OUTPUT_DIR}.")


if __name__ == "__main__":
    main()
